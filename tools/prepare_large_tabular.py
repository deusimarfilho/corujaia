from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Iterable, Iterator

import pandas as pd
from openpyxl import load_workbook


def safe_name(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9._-]+", "_", value)
    return value.strip("_") or "sheet"


def iter_csv_rows(path: Path) -> Iterator[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        yield from reader


def iter_xlsx_sheets(path: Path) -> Iterator[tuple[str, Iterator[list[str]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        def row_iter() -> Iterator[list[str]]:
            for row in sheet.iter_rows(values_only=True):
                yield ["" if value is None else str(value) for value in row]

        yield sheet.title, row_iter()


def iter_xls_sheets(path: Path) -> Iterator[tuple[str, Iterator[list[str]]]]:
    workbook = pd.ExcelFile(path, engine="xlrd")
    for sheet_name in workbook.sheet_names:
        dataframe = workbook.parse(sheet_name=sheet_name, dtype=str).fillna("")

        def row_iter() -> Iterator[list[str]]:
            yield [str(column) for column in dataframe.columns.tolist()]
            for row in dataframe.itertuples(index=False, name=None):
                yield ["" if value is None else str(value) for value in row]

        yield sheet_name, row_iter()


def chunk_rows(
    rows: Iterable[list[str]],
    output_prefix: Path,
    rows_per_file: int,
) -> tuple[int, int]:
    iterator = iter(rows)
    header = next(iterator, None)
    if header is None:
        return 0, 0

    total_rows = 0
    file_count = 0
    current_chunk: list[list[str]] = []

    def flush() -> None:
        nonlocal current_chunk, file_count
        if not current_chunk:
            return
        file_count += 1
        output_path = output_prefix.parent / f"{output_prefix.name}_part_{file_count:03d}.csv"
        with output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(header)
            writer.writerows(current_chunk)
        current_chunk = []

    for row in iterator:
        current_chunk.append(row)
        total_rows += 1
        if len(current_chunk) >= rows_per_file:
            flush()

    flush()
    return total_rows, file_count


def process_csv(path: Path, output_dir: Path, rows_per_file: int) -> dict:
    output_prefix = output_dir / safe_name(path.stem)
    total_rows, file_count = chunk_rows(iter_csv_rows(path), output_prefix, rows_per_file)
    return {
        "source": str(path),
        "type": "csv",
        "rows_per_file": rows_per_file,
        "total_rows": total_rows,
        "files_created": file_count,
    }


def process_xlsx(path: Path, output_dir: Path, rows_per_file: int) -> dict:
    summary = {
        "source": str(path),
        "type": "xlsx",
        "rows_per_file": rows_per_file,
        "sheets": [],
    }

    for sheet_name, rows in iter_xlsx_sheets(path):
        sheet_dir = output_dir / safe_name(path.stem)
        sheet_prefix = sheet_dir / safe_name(sheet_name)
        total_rows, file_count = chunk_rows(rows, sheet_prefix, rows_per_file)
        summary["sheets"].append(
            {
                "sheet": sheet_name,
                "total_rows": total_rows,
                "files_created": file_count,
            }
        )

    return summary


def process_xls(path: Path, output_dir: Path, rows_per_file: int) -> dict:
    summary = {
        "source": str(path),
        "type": "xls",
        "rows_per_file": rows_per_file,
        "sheets": [],
    }

    for sheet_name, rows in iter_xls_sheets(path):
        sheet_dir = output_dir / safe_name(path.stem)
        sheet_prefix = sheet_dir / safe_name(sheet_name)
        total_rows, file_count = chunk_rows(rows, sheet_prefix, rows_per_file)
        summary["sheets"].append(
            {
                "sheet": sheet_name,
                "total_rows": total_rows,
                "files_created": file_count,
            }
        )

    return summary


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Quebra arquivos tabulares grandes em partes menores para upload no AnythingLLM."
    )
    parser.add_argument("input_file", help="Arquivo .xls, .xlsx ou .csv de entrada.")
    parser.add_argument(
        "--rows-per-file",
        type=int,
        default=2000,
        help="Quantidade maxima de linhas de dados por arquivo gerado. Padrao: 2000.",
    )
    parser.add_argument(
        "--output-dir",
        default="prepared_uploads",
        help="Pasta de saida relativa ao projeto ou absoluta. Padrao: prepared_uploads",
    )
    args = parser.parse_args()

    input_path = Path(args.input_file).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    if not input_path.exists():
        raise SystemExit(f"Arquivo nao encontrado: {input_path}")

    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        summary = process_csv(input_path, output_dir, args.rows_per_file)
    elif suffix == ".xls":
        summary = process_xls(input_path, output_dir, args.rows_per_file)
    elif suffix == ".xlsx":
        summary = process_xlsx(input_path, output_dir, args.rows_per_file)
    else:
        raise SystemExit("Formato nao suportado. Use .xls, .xlsx ou .csv.")

    manifest_path = output_dir / f"{safe_name(input_path.stem)}_manifest.json"
    manifest_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    print(f"\nManifest salvo em: {manifest_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
